#!/usr/bin/env python3
"""
recalc.py — Excel formula recalculation and error detection.

Opens an openpyxl-generated .xlsx file, scans all cells for formula error
values (#REF!, #DIV/0!, #VALUE!, #NAME?, #NULL!, #N/A, #NUM!), and reports
results as JSON.

Usage:
    python recalc.py <excel_file> [timeout_seconds]

Output (JSON):
    {
      "status": "success" | "errors_found",
      "total_errors": int,
      "total_formulas": int,
      "error_summary": {
          "#REF!": {"count": N, "locations": ["Sheet!A1", ...]},
          ...
      }
    }
"""

import json
import signal
import sys
from pathlib import Path

EXCEL_ERRORS = {
    "#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#N/A", "#NUM!"
}


def recalc(excel_path: str) -> dict:
    """Scan an Excel file for formula errors and return a structured report."""
    try:
        import openpyxl
    except ImportError:
        return {
            "status": "errors_found",
            "total_errors": 1,
            "total_formulas": 0,
            "error_summary": {
                "ImportError": {
                    "count": 1,
                    "locations": ["openpyxl not installed. Run: pip install openpyxl"],
                }
            },
        }

    path = Path(excel_path)
    if not path.exists():
        print(f"Error: File not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    # Load with data_only=False to see formulas, and data_only=True to see
    # cached values (which is where Excel stores error results).
    wb_formulas = openpyxl.load_workbook(str(path), data_only=False)
    wb_values = openpyxl.load_workbook(str(path), data_only=True)

    total_formulas = 0
    error_summary: dict[str, dict] = {}

    for sheet_name in wb_formulas.sheetnames:
        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]

        for row in ws_formulas.iter_rows():
            for cell in row:
                # Check if cell contains a formula
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1

                    # Get the cached value from the data_only workbook
                    val_cell = ws_values[cell.coordinate]
                    cached = val_cell.value

                    # Check for error values in cached result
                    if isinstance(cached, str) and cached in EXCEL_ERRORS:
                        loc = f"{sheet_name}!{cell.coordinate}"
                        if cached not in error_summary:
                            error_summary[cached] = {"count": 0, "locations": []}
                        error_summary[cached]["count"] += 1
                        error_summary[cached]["locations"].append(loc)

    total_errors = sum(e["count"] for e in error_summary.values())

    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": total_formulas,
    }
    if total_errors > 0:
        result["error_summary"] = error_summary

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]", file=sys.stderr)
        sys.exit(1)

    excel_path = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    # Set timeout (Unix only; silently skip on Windows)
    if hasattr(signal, "SIGALRM"):
        def _timeout_handler(signum, frame):
            print(
                json.dumps(
                    {
                        "status": "errors_found",
                        "total_errors": 1,
                        "total_formulas": 0,
                        "error_summary": {
                            "Timeout": {
                                "count": 1,
                                "locations": [f"Exceeded {timeout}s limit"],
                            }
                        },
                    }
                )
            )
            sys.exit(1)

        signal.signal(signal.SIGALRM, _timeout_handler)
        signal.alarm(timeout)

    result = recalc(excel_path)
    print(json.dumps(result, indent=2))

    if hasattr(signal, "SIGALRM"):
        signal.alarm(0)

    sys.exit(0 if result["status"] == "success" else 1)


if __name__ == "__main__":
    main()
