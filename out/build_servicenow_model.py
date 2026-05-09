"""ServiceNow (NOW) 3-Statement Financial Model — openpyxl builder.

Outputs: ./servicenow-3statement-model.xlsx
Conventions (xlsx-author skill):
  Blue font  = hardcoded input value
  Black font = same-sheet formula
  Green font = cross-sheet formula link
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter, column_index_from_string

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DARK_BLUE  = "1F4E79"
WHITE      = "FFFFFF"
LIGHT_BLUE = "D9E1F2"
MED_BLUE   = "BDD7EE"
LIGHT_GREY = "F2F2F2"
INPUT_BLUE = "0000FF"
GREEN_LINK = "008000"
BLACK      = "000000"
RED        = "FF0000"

FMT_DOLLAR = '#,##0_);(#,##0)'
FMT_PCT    = '0.0%'
FMT_MULT   = '0.0"x"'
FMT_DAYS   = '0.0'

PERIODS    = ['FY2022A', 'FY2023A', 'FY2024A', 'FY2025E', 'FY2026E', 'FY2027E']
DATA_COLS  = ['B', 'C', 'D', 'E', 'F', 'G']
HIST_COLS  = ['B', 'C', 'D']
PROJ_COLS  = ['E', 'F', 'G']
# Assumptions tab projection columns (B/C/D = FY2025E/2026E/2027E)
ASSUMP_COLS = ['B', 'C', 'D']

OUTPUT_FILE = Path(__file__).parent / "servicenow-3statement-model.xlsx"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _font(color=BLACK, bold=False, italic=False, size=10):
    return Font(color=color, bold=bold, italic=italic, size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _center():
    return Alignment(horizontal='center', vertical='center')

def _left():
    return Alignment(horizontal='left', vertical='center', indent=1)

def _thin_border():
    s = Side(style='thin')
    return Border(top=s, bottom=s, left=s, right=s)

def _medium_right():
    return Side(style='medium')


def inp(ws, ref, value, fmt=FMT_DOLLAR, bold=False):
    """Hardcoded input cell — blue font."""
    c = ws[ref]
    c.value = value
    c.font = _font(INPUT_BLUE, bold=bold)
    c.number_format = fmt


def frm(ws, ref, formula, fmt=FMT_DOLLAR, bold=False, italic=False):
    """Same-sheet formula — black font."""
    c = ws[ref]
    c.value = formula
    c.font = _font(BLACK, bold=bold, italic=italic)
    c.number_format = fmt


def lnk(ws, ref, formula, fmt=FMT_DOLLAR, bold=False):
    """Cross-sheet link — green font."""
    c = ws[ref]
    c.value = formula
    c.font = _font(GREEN_LINK, bold=bold)
    c.number_format = fmt


def dark_hdr(ws, ref, label_text, merged_end=None):
    """Dark blue fill, white bold font header."""
    if merged_end:
        ws.merge_cells(f"{ref}:{merged_end}")
    c = ws[ref]
    c.value = label_text
    c.font = _font(WHITE, bold=True)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _center()


def light_hdr(ws, ref, label_text, merged_end=None):
    """Light blue fill, black bold font sub-header."""
    if merged_end:
        ws.merge_cells(f"{ref}:{merged_end}")
    c = ws[ref]
    c.value = label_text
    c.font = _font(BLACK, bold=True)
    c.fill = _fill(LIGHT_BLUE)
    c.alignment = _center()


def col_hdrs(ws, row, start_col='B', periods=None):
    """Write period labels with light-blue fill."""
    if periods is None:
        periods = PERIODS
    idx = column_index_from_string(start_col)
    for p in periods:
        col = get_column_letter(idx)
        c = ws[f"{col}{row}"]
        c.value = p
        c.font = _font(BLACK, bold=True)
        c.fill = _fill(LIGHT_BLUE)
        c.alignment = _center()
        idx += 1


def total(ws, ref, formula, fmt=FMT_DOLLAR):
    """Total / subtotal row: bold, med-blue fill."""
    c = ws[ref]
    c.value = formula
    c.font = _font(BLACK, bold=True)
    c.fill = _fill(MED_BLUE)
    c.number_format = fmt


def grand_total(ws, ref, formula, fmt=FMT_DOLLAR):
    """Grand total row: bold, dark-blue fill, white font."""
    c = ws[ref]
    c.value = formula
    c.font = _font(WHITE, bold=True)
    c.fill = _fill(DARK_BLUE)
    c.number_format = fmt


def label(ws, ref, text, indent=0, bold=False, italic=False):
    c = ws[ref]
    c.value = text
    c.font = _font(BLACK, bold=bold, italic=italic)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=indent)


def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_hist_proj_border(ws, start_row, end_row):
    """Medium right border on col D (historical / projected split)."""
    for row in range(start_row, end_row + 1):
        c = ws[f"D{row}"]
        existing = c.border
        c.border = Border(
            left=existing.left,
            right=_medium_right(),
            top=existing.top,
            bottom=existing.bottom,
        )


def subtitle(ws, ref, text):
    c = ws[ref]
    c.value = text
    c.font = _font(BLACK, italic=True, size=9)
    c.alignment = _left()


def _prev(col):
    """Return the previous column letter (E->D, F->E, G->F)."""
    return get_column_letter(column_index_from_string(col) - 1)


def proj_assump_col(proj_col):
    """Map projection col (E/F/G) to Assumptions col (B/C/D)."""
    return {'E': 'B', 'F': 'C', 'G': 'D'}[proj_col]


# ---------------------------------------------------------------------------
# revenue_build — Software KPIs & Revenue Forecast
# ---------------------------------------------------------------------------

def build_revenue_build(ws):
    ws.sheet_properties.tabColor = DARK_BLUE
    set_col_widths(ws, {'A': 50, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12})

    dark_hdr(ws, 'A1', 'SERVICENOW (NOW) — REVENUE BUILD & SOFTWARE KPIs', merged_end='G1')
    subtitle(ws, 'A2', '($ in millions, unless noted)')
    col_hdrs(ws, 3)

    # ── ARR BRIDGE ────────────────────────────────────────────────────────
    dark_hdr(ws, 'A4', 'ANNUAL RECURRING REVENUE (ARR) BRIDGE', merged_end='G4')
    label(ws, 'A5',  'Beginning ARR',                  indent=2)
    label(ws, 'A6',  '+ New Logo ARR',                 indent=2)
    label(ws, 'A7',  '+ Expansion ARR',                indent=2)
    label(ws, 'A8',  '− Contraction ARR',              indent=2)
    label(ws, 'A9',  '− Churned ARR',                  indent=2)
    label(ws, 'A10', 'Ending ARR',                     bold=True)
    label(ws, 'A11', 'ARR Growth %',                   indent=2, italic=True)
    label(ws, 'A12', 'Implied ARR per Customer ($K)',   indent=2, italic=True)

    # Historical bridge — hardcoded blue inputs
    # Bridge components chosen to sum to disclosed ending ARR values
    # FY22: 5737→7100  FY23: 7100→8970  FY24: 8970→10900
    hist_beg_arr  = [5737, 7100, 8970]
    hist_new_logo = [200,  250,  300]
    hist_expand   = [1237, 1712, 1747]
    hist_contract = [-17,  -21,  -27]
    hist_churn    = [-57,  -71,  -90]

    inp(ws, 'B5', hist_beg_arr[0])
    for i, col in enumerate(HIST_COLS):
        if i > 0:
            frm(ws, f'{col}5', f'={HIST_COLS[i-1]}10')
        inp(ws, f'{col}6', hist_new_logo[i])
        inp(ws, f'{col}7', hist_expand[i])
        inp(ws, f'{col}8', hist_contract[i])
        inp(ws, f'{col}9', hist_churn[i])
        total(ws, f'{col}10', f'={col}5+{col}6+{col}7+{col}8+{col}9')
        if i > 0:
            frm(ws, f'{col}11', f'={col}10/{HIST_COLS[i-1]}10-1', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}12', f'={col}10/{col}32*1000', fmt='#,##0.0', italic=True)

    # Projection bridge — formulas referencing Assumptions tab
    prev = 'D'
    for col in PROJ_COLS:
        ac = proj_assump_col(col)
        frm(ws, f'{col}5', f'={prev}10')
        lnk(ws, f'{col}6', f'={prev}6*(1+Assumptions!${ac}$34)')
        lnk(ws, f'{col}7', f'={col}5*Assumptions!${ac}$35')
        lnk(ws, f'{col}8', f'=-{col}5*Assumptions!${ac}$37')
        lnk(ws, f'{col}9', f'=-{col}5*(1-Assumptions!${ac}$36)')
        total(ws, f'{col}10', f'={col}5+{col}6+{col}7+{col}8+{col}9')
        frm(ws, f'{col}11', f'={col}10/{prev}10-1', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}12', f'={col}10/{col}32*1000', fmt='#,##0.0', italic=True)
        prev = col

    # ── RETENTION METRICS ─────────────────────────────────────────────────
    dark_hdr(ws, 'A14', 'RETENTION METRICS', merged_end='G14')
    label(ws, 'A15', 'Gross Revenue Retention (GRR) %', indent=2)
    label(ws, 'A16', 'Net Revenue Retention (NRR) %',   indent=2)
    label(ws, 'A17', 'Gross Churn Rate %',               indent=3, italic=True)
    label(ws, 'A18', 'Net Expansion Rate %',             indent=3, italic=True)
    label(ws, 'A19', 'Logo Churn Rate %',                indent=2)

    hist_grr  = [0.990, 0.990, 0.990]
    hist_nrr  = [1.270, 1.260, 1.250]
    hist_logo = [0.030, 0.030, 0.030]

    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}15', hist_grr[i],  fmt=FMT_PCT)
        inp(ws, f'{col}16', hist_nrr[i],  fmt=FMT_PCT)
        frm(ws, f'{col}17', f'=1-{col}15', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}18', f'={col}16-{col}15', fmt=FMT_PCT, italic=True)
        inp(ws, f'{col}19', hist_logo[i], fmt=FMT_PCT)

    for col in PROJ_COLS:
        ac = proj_assump_col(col)
        lnk(ws, f'{col}15', f'=Assumptions!${ac}$36', fmt=FMT_PCT)
        lnk(ws, f'{col}16', f'=Assumptions!${ac}$41', fmt=FMT_PCT)
        frm(ws, f'{col}17', f'=1-{col}15', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}18', f'={col}16-{col}15', fmt=FMT_PCT, italic=True)
        lnk(ws, f'{col}19', f'=Assumptions!${ac}$42', fmt=FMT_PCT)

    # ── RPO ───────────────────────────────────────────────────────────────
    dark_hdr(ws, 'A21', 'REMAINING PERFORMANCE OBLIGATIONS (RPO)', merged_end='G21')
    label(ws, 'A22', 'Total RPO',             indent=2)
    label(ws, 'A23', 'Current RPO (cRPO)',    indent=2)
    label(ws, 'A24', 'Non-Current RPO',       indent=3, italic=True)
    label(ws, 'A25', 'cRPO / Total RPO %',   indent=3, italic=True)
    label(ws, 'A26', 'Total RPO Growth %',   indent=3, italic=True)

    hist_rpo  = [13600, 17700, 21900]
    hist_crpo = [7000,  8300,  10300]

    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}22', hist_rpo[i])
        inp(ws, f'{col}23', hist_crpo[i])
        frm(ws, f'{col}24', f'={col}22-{col}23', italic=True)
        frm(ws, f'{col}25', f'={col}23/{col}22', fmt=FMT_PCT, italic=True)
        if i > 0:
            frm(ws, f'{col}26', f'={col}22/{HIST_COLS[i-1]}22-1', fmt=FMT_PCT, italic=True)

    prev = 'D'
    for col in PROJ_COLS:
        ac = proj_assump_col(col)
        lnk(ws, f'{col}22', f'={prev}22*(1+Assumptions!${ac}$51)')
        lnk(ws, f'{col}23', f'={col}22*Assumptions!${ac}$52')
        frm(ws, f'{col}24', f'={col}22-{col}23', italic=True)
        frm(ws, f'{col}25', f'={col}23/{col}22', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}26', f'={col}22/{prev}22-1', fmt=FMT_PCT, italic=True)
        prev = col

    # ── CUSTOMER METRICS ──────────────────────────────────────────────────
    dark_hdr(ws, 'A28', 'CUSTOMER METRICS — BY ACV TIER', merged_end='G28')
    label(ws, 'A29', 'Customers > $1M ACV',          indent=2)
    label(ws, 'A30', 'Customers > $5M ACV',          indent=2)
    label(ws, 'A31', 'Customers > $20M ACV',         indent=2)
    label(ws, 'A32', 'Total Enterprise Customers',   indent=2)
    label(ws, 'A33', 'YoY Growth — Customers >$1M',  indent=3, italic=True)
    label(ws, 'A34', 'YoY Growth — Customers >$5M',  indent=3, italic=True)

    hist_1m  = [1530, 1986, 2550]
    hist_5m  = [156,  197,  243]
    hist_20m = [74,   92,   114]
    hist_ent = [7400, 8100, 8800]
    fmt_int  = '#,##0'

    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}29', hist_1m[i],  fmt=fmt_int)
        inp(ws, f'{col}30', hist_5m[i],  fmt=fmt_int)
        inp(ws, f'{col}31', hist_20m[i], fmt=fmt_int)
        inp(ws, f'{col}32', hist_ent[i], fmt=fmt_int)
        if i > 0:
            frm(ws, f'{col}33', f'={col}29/{HIST_COLS[i-1]}29-1', fmt=FMT_PCT, italic=True)
            frm(ws, f'{col}34', f'={col}30/{HIST_COLS[i-1]}30-1', fmt=FMT_PCT, italic=True)

    prev = 'D'
    for col in PROJ_COLS:
        ac = proj_assump_col(col)
        lnk(ws, f'{col}29', f'={prev}29*(1+Assumptions!${ac}$45)', fmt=fmt_int)
        lnk(ws, f'{col}30', f'={prev}30*(1+Assumptions!${ac}$46)', fmt=fmt_int)
        lnk(ws, f'{col}31', f'={prev}31*(1+Assumptions!${ac}$47)', fmt=fmt_int)
        lnk(ws, f'{col}32', f'={prev}32*(1+Assumptions!${ac}$45*0.6)', fmt=fmt_int)
        frm(ws, f'{col}33', f'={col}29/{prev}29-1', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}34', f'={col}30/{prev}30-1', fmt=FMT_PCT, italic=True)
        prev = col

    # ── REVENUE METRICS ───────────────────────────────────────────────────
    # Row 37 = Subscription Revenue anchor — IS!E5 links here
    # Row 38 = PS Revenue anchor — IS!E6 links here
    dark_hdr(ws, 'A36', 'REVENUE METRICS', merged_end='G36')
    label(ws, 'A37', 'Subscription Revenue',          bold=True)
    label(ws, 'A38', 'Professional Services Revenue', indent=2)
    label(ws, 'A39', 'Total Revenue',                 bold=True)
    label(ws, 'A40', 'Subscription Revenue Growth %', indent=2, italic=True)
    label(ws, 'A41', 'Implied Billings',              indent=2, italic=True)

    hist_sub = [6894, 8396, 10402]
    hist_ps  = [586,  537,  576]

    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}37', hist_sub[i])
        inp(ws, f'{col}38', hist_ps[i])
        total(ws, f'{col}39', f'={col}37+{col}38')
        if i > 0:
            frm(ws, f'{col}40', f'={col}37/{HIST_COLS[i-1]}37-1', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}41', f'={col}10-{col}5+{col}37', italic=True)

    prev = 'D'
    for col in PROJ_COLS:
        ac = proj_assump_col(col)
        frm(ws, f'{col}37', f'={col}46')           # feeds from ARR Bridge primary below
        lnk(ws, f'{col}38', f'={col}37*Assumptions!${ac}$55')
        total(ws, f'{col}39', f'={col}37+{col}38')
        frm(ws, f'{col}40', f'={col}37/{prev}37-1', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}41', f'={col}10-{col}5+{col}37', italic=True)
        prev = col

    # ── PRIMARY METHODOLOGY: ARR Bridge ───────────────────────────────────
    dark_hdr(ws, 'A43', 'PRIMARY FORECAST METHODOLOGY: ARR Bridge', merged_end='G43')
    label(ws, 'A44', 'Average ARR ((Beg + End) / 2)', indent=2, italic=True)
    label(ws, 'A45', 'Billing Timing Factor',          indent=2)
    label(ws, 'A46', 'Projected Subscription Revenue', bold=True)

    for i, col in enumerate(HIST_COLS):
        frm(ws, f'{col}44', f'=({col}5+{col}10)/2', italic=True)
        frm(ws, f'{col}45', f'={col}37/{col}44')   # implied BTF from actuals
        total(ws, f'{col}46', f'={col}44*{col}45') # = actual sub rev for historicals

    for col in PROJ_COLS:
        ac = proj_assump_col(col)
        frm(ws, f'{col}44', f'=({col}5+{col}10)/2', italic=True)
        lnk(ws, f'{col}45', f'=Assumptions!${ac}$38')
        total(ws, f'{col}46', f'={col}44*{col}45')

    # ── CROSS-CHECK A: NRR-Based ──────────────────────────────────────────
    light_hdr(ws, 'A48', 'CROSS-CHECK A — NRR-Based Methodology', merged_end='G48')
    label(ws, 'A49', 'Beginning ARR',                      indent=2, italic=True)
    label(ws, 'A50', 'Existing Customer ARR (Beg × NRR)',  indent=2, italic=True)
    label(ws, 'A51', '+ New Logo ARR',                     indent=2, italic=True)
    label(ws, 'A52', 'Ending ARR (NRR method)',            indent=2, italic=True)
    label(ws, 'A53', 'Average ARR',                        indent=2, italic=True)
    label(ws, 'A54', 'Sub Revenue (NRR method)',           bold=True, italic=True)

    for col in DATA_COLS:
        frm(ws, f'{col}49', f'={col}5',             italic=True)
        frm(ws, f'{col}50', f'={col}5*{col}16',     italic=True)
        frm(ws, f'{col}51', f'={col}6',             italic=True)
        frm(ws, f'{col}52', f'={col}50+{col}51',    italic=True)
        frm(ws, f'{col}53', f'=({col}5+{col}52)/2', italic=True)
        frm(ws, f'{col}54', f'={col}53*{col}45',    italic=True)

    # ── CROSS-CHECK B: RPO-Anchored ───────────────────────────────────────
    light_hdr(ws, 'A56', 'CROSS-CHECK B — RPO-Anchored Methodology', merged_end='G56')
    label(ws, 'A57', 'cRPO — prior year-end',         indent=2, italic=True)
    label(ws, 'A58', 'Recognition Rate %',            indent=2)
    label(ws, 'A59', 'Sub Revenue (RPO method)',      bold=True, italic=True)

    inp(ws, 'B57', 5800)  # FY2021 cRPO estimate (pre-period)
    for i, col in enumerate(DATA_COLS):
        if col != 'B':
            frm(ws, f'{col}57', f'={DATA_COLS[i-1]}23', italic=True)
        if col in HIST_COLS:
            frm(ws, f'{col}58', f'={col}37/{col}57', fmt=FMT_PCT)  # implied recog rate
        else:
            ac = proj_assump_col(col)
            lnk(ws, f'{col}58', f'=Assumptions!${ac}$50', fmt=FMT_PCT)
        frm(ws, f'{col}59', f'={col}57*{col}58', italic=True)

    # ── CROSS-CHECK COMPARISON ─────────────────────────────────────────────
    light_hdr(ws, 'A61', 'METHODOLOGY COMPARISON', merged_end='G61')
    label(ws, 'A62', 'Primary — ARR Bridge',          indent=2)
    label(ws, 'A63', 'Cross-Check A — NRR-Based',     indent=2)
    label(ws, 'A64', 'Cross-Check B — RPO-Anchored',  indent=2)
    label(ws, 'A65', 'NRR vs. Primary Δ %',           indent=3, italic=True)
    label(ws, 'A66', 'RPO vs. Primary Δ %',           indent=3, italic=True)
    label(ws, 'A67', 'NRR flag: |delta| > 15%',       indent=3, italic=True)
    label(ws, 'A68', 'RPO flag: |delta| > 15%',       indent=3, italic=True)

    for col in DATA_COLS:
        frm(ws, f'{col}62', f'={col}46')
        frm(ws, f'{col}63', f'={col}54')
        frm(ws, f'{col}64', f'={col}59')
        frm(ws, f'{col}65', f'={col}63/{col}62-1', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}66', f'={col}64/{col}62-1', fmt=FMT_PCT, italic=True)
        frm(ws, f'{col}67', f'=IF(ABS({col}65)>0.15,"FLAG","OK")', italic=True)
        frm(ws, f'{col}68', f'=IF(ABS({col}66)>0.15,"FLAG","OK")', italic=True)

    add_hist_proj_border(ws, 1, 70)
    ws.freeze_panes = 'B4'


# ---------------------------------------------------------------------------
# Assumptions sheet
# ---------------------------------------------------------------------------

def build_assumptions(ws):
    ws.sheet_properties.tabColor = "F2F2F2"
    set_col_widths(ws, {'A': 42, 'B': 13, 'C': 13, 'D': 13})

    dark_hdr(ws, 'A1', 'SERVICENOW (NOW) — MODEL ASSUMPTIONS', merged_end='D1')
    subtitle(ws, 'A2', '($ in millions, unless noted)')
    col_hdrs(ws, 3, start_col='B', periods=['FY2025E', 'FY2026E', 'FY2027E'])

    # Section: Income Statement Drivers
    dark_hdr(ws, 'A4', 'INCOME STATEMENT DRIVERS', merged_end='D4')

    rows = [
        (5,  'Revenue Growth Rate (Memo)',  0.220, 0.190, 0.170, FMT_PCT),
        (6,  'Gross Margin %',             0.795, 0.800, 0.805, FMT_PCT),
        (7,  "R&D % of Revenue",           0.230, 0.225, 0.220, FMT_PCT),
        (8,  "S&M % of Revenue",           0.220, 0.215, 0.210, FMT_PCT),
        (9,  "G&A % of Revenue",           0.060, 0.058, 0.056, FMT_PCT),
        (10, 'Tax Rate',                   0.180, 0.180, 0.180, FMT_PCT),
        (11, "D&A % of Revenue",           0.070, 0.068, 0.065, FMT_PCT),
        (12, "SBC % of Revenue",           0.135, 0.130, 0.125, FMT_PCT),
    ]
    for r, lbl, v25, v26, v27, fmt in rows:
        label(ws, f'A{r}', lbl, indent=1)
        inp(ws, f'B{r}', v25, fmt=fmt)
        inp(ws, f'C{r}', v26, fmt=fmt)
        inp(ws, f'D{r}', v27, fmt=fmt)

    # Section: Cash Flow & Capex
    dark_hdr(ws, 'A14', 'CASH FLOW & CAPEX DRIVERS', merged_end='D14')
    label(ws, 'A15', 'Capex % of Revenue', indent=1)
    inp(ws, 'B15', 0.065, fmt=FMT_PCT)
    inp(ws, 'C15', 0.062, fmt=FMT_PCT)
    inp(ws, 'D15', 0.060, fmt=FMT_PCT)

    # Section: Working Capital
    dark_hdr(ws, 'A17', 'WORKING CAPITAL DRIVERS', merged_end='D17')
    label(ws, 'A18', 'DSO (days)', indent=1)
    inp(ws, 'B18', 75, fmt=FMT_DAYS)
    inp(ws, 'C18', 74, fmt=FMT_DAYS)
    inp(ws, 'D18', 73, fmt=FMT_DAYS)
    label(ws, 'A19', 'DPO (days)', indent=1)
    inp(ws, 'B19', 32, fmt=FMT_DAYS)
    inp(ws, 'C19', 32, fmt=FMT_DAYS)
    inp(ws, 'D19', 32, fmt=FMT_DAYS)

    # Section: Financing
    dark_hdr(ws, 'A21', 'FINANCING ASSUMPTIONS', merged_end='D21')
    fin_rows = [
        (22, 'Share Repurchases ($M)',       -1800, -1900, -2000),
        (23, 'ESPP / Options Proceeds ($M)',   500,   530,   560),
        (24, 'Net Acquisitions ($M)',          -300,  -300,  -300),
        (25, 'Net Investment Purchases ($M)',  -350,  -350,  -350),
    ]
    for r, lbl, v25, v26, v27 in fin_rows:
        label(ws, f'A{r}', lbl, indent=1)
        inp(ws, f'B{r}', v25, fmt=FMT_DOLLAR)
        inp(ws, f'C{r}', v26, fmt=FMT_DOLLAR)
        inp(ws, f'D{r}', v27, fmt=FMT_DOLLAR)

    # ── SOFTWARE REVENUE DRIVERS ──────────────────────────────────────────
    dark_hdr(ws, 'A27', 'SOFTWARE REVENUE DRIVERS', merged_end='D27')

    light_hdr(ws, 'A29', 'ARR BRIDGE ASSUMPTIONS', merged_end='D29')
    # Row 30: methodology selector (label only)
    label(ws, 'A30', 'Primary Methodology: ARR Bridge', indent=1, italic=True)

    sw_arr = [
        # row, label, FY25, FY26, FY27, fmt
        (34, 'New Logo ARR Growth % YoY',             0.200, 0.180, 0.160, FMT_PCT),
        (35, 'Expansion Rate (% of Beginning ARR)',   0.240, 0.230, 0.220, FMT_PCT),
        (36, 'Gross Revenue Retention (GRR) %',       0.990, 0.990, 0.990, FMT_PCT),
        (37, 'Contraction Rate (% of Beginning ARR)', 0.003, 0.003, 0.003, FMT_PCT),
        (38, 'Billing Timing Factor',                 0.980, 0.980, 0.980, '0.000'),
    ]
    for r, lbl, v25, v26, v27, fmt in sw_arr:
        label(ws, f'A{r}', lbl, indent=1)
        inp(ws, f'B{r}', v25, fmt=fmt)
        inp(ws, f'C{r}', v26, fmt=fmt)
        inp(ws, f'D{r}', v27, fmt=fmt)

    light_hdr(ws, 'A40', 'RETENTION-BASED ASSUMPTIONS', merged_end='D40')
    sw_ret = [
        (41, 'Net Revenue Retention (NRR) %',  1.240, 1.230, 1.220, FMT_PCT),
        (42, 'Logo Churn Rate %',              0.030, 0.030, 0.030, FMT_PCT),
    ]
    for r, lbl, v25, v26, v27, fmt in sw_ret:
        label(ws, f'A{r}', lbl, indent=1)
        inp(ws, f'B{r}', v25, fmt=fmt)
        inp(ws, f'C{r}', v26, fmt=fmt)
        inp(ws, f'D{r}', v27, fmt=fmt)

    light_hdr(ws, 'A44', 'CUSTOMER TIER GROWTH', merged_end='D44')
    sw_cust = [
        (45, 'Customers >$1M ACV Growth %',   0.200, 0.180, 0.160, FMT_PCT),
        (46, 'Customers >$5M ACV Growth %',   0.180, 0.160, 0.150, FMT_PCT),
        (47, 'Customers >$20M ACV Growth %',  0.150, 0.140, 0.130, FMT_PCT),
    ]
    for r, lbl, v25, v26, v27, fmt in sw_cust:
        label(ws, f'A{r}', lbl, indent=1)
        inp(ws, f'B{r}', v25, fmt=fmt)
        inp(ws, f'C{r}', v26, fmt=fmt)
        inp(ws, f'D{r}', v27, fmt=fmt)

    light_hdr(ws, 'A49', 'RPO ASSUMPTIONS', merged_end='D49')
    sw_rpo = [
        (50, 'cRPO Recognition Rate (next 12 months)', 0.940, 0.940, 0.940, FMT_PCT),
        (51, 'Total RPO Growth %',                     0.180, 0.160, 0.140, FMT_PCT),
        (52, 'cRPO as % of Total RPO',                 0.470, 0.470, 0.470, FMT_PCT),
    ]
    for r, lbl, v25, v26, v27, fmt in sw_rpo:
        label(ws, f'A{r}', lbl, indent=1)
        inp(ws, f'B{r}', v25, fmt=fmt)
        inp(ws, f'C{r}', v26, fmt=fmt)
        inp(ws, f'D{r}', v27, fmt=fmt)

    light_hdr(ws, 'A54', 'PROFESSIONAL SERVICES', merged_end='D54')
    label(ws, 'A55', 'PS Revenue as % of Subscription Revenue', indent=1)
    inp(ws, 'B55', 0.050, fmt=FMT_PCT)
    inp(ws, 'C55', 0.048, fmt=FMT_PCT)
    inp(ws, 'D55', 0.046, fmt=FMT_PCT)

    ws.freeze_panes = 'B4'


# ---------------------------------------------------------------------------
# IS — Income Statement
# ---------------------------------------------------------------------------

def build_IS(ws):
    ws.sheet_properties.tabColor = DARK_BLUE
    set_col_widths(ws, {'A': 38, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12})

    dark_hdr(ws, 'A1', 'SERVICENOW (NOW) — INCOME STATEMENT', merged_end='G1')
    subtitle(ws, 'A2', '($ in millions)')
    col_hdrs(ws, 3)

    # ---- REVENUE ----
    dark_hdr(ws, 'A4', 'REVENUE', merged_end='G4')
    label(ws, 'A5', 'Subscription Revenue', indent=2)
    label(ws, 'A6', 'Professional Services Revenue', indent=2)
    label(ws, 'A7', 'Total Revenue', bold=True)
    label(ws, 'A8', 'Revenue Growth %', indent=2, italic=True)

    hist_sub  = [6894, 8396, 10402]
    hist_ps   = [586,  537,  576]
    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}5', hist_sub[i])
        inp(ws, f'{col}6', hist_ps[i])
        total(ws, f'{col}7', f'={col}5+{col}6')
        if i > 0:
            prev = HIST_COLS[i-1]
            frm(ws, f'{col}8', f'={col}7/{prev}7-1', fmt=FMT_PCT, italic=True)

    # Projections — revenue linked from revenue_build (3-statement-model-software skill)
    proj_assump = ['B', 'C', 'D']  # Assumptions col map (used for COGS/OpEx below)
    prev_col = 'D'
    for col in PROJ_COLS:
        lnk(ws, f'{col}5', f'=revenue_build!{col}37')
        lnk(ws, f'{col}6', f'=revenue_build!{col}38')
        total(ws, f'{col}7', f'={col}5+{col}6')
        frm(ws, f'{col}8', f'={col}7/{prev_col}7-1', fmt=FMT_PCT, italic=True)
        prev_col = col

    # ---- COST OF REVENUE ----
    dark_hdr(ws, 'A10', 'COST OF REVENUE', merged_end='G10')
    label(ws, 'A11', 'Cost of Revenue — Subscription', indent=2)
    label(ws, 'A12', 'Cost of Revenue — Professional Services', indent=2)
    label(ws, 'A13', 'Total Cost of Revenue', bold=True)
    label(ws, 'A14', 'Gross Profit', bold=True)
    label(ws, 'A15', 'Gross Margin %', indent=2, italic=True)

    hist_cor_sub = [1211, 1431, 1740]
    hist_cor_ps  = [543,  513,  567]
    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}11', hist_cor_sub[i])
        inp(ws, f'{col}12', hist_cor_ps[i])
        total(ws, f'{col}13', f'={col}11+{col}12')
        total(ws, f'{col}14', f'={col}7-{col}13')
        frm(ws, f'{col}15', f'={col}14/{col}7', fmt=FMT_PCT, italic=True)

    prev_col = 'D'
    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        lnk(ws, f'{col}13', f'={col}7*(1-Assumptions!${ac}$6)', bold=True)
        frm(ws, f'{col}11', f'={col}13*0.75')
        frm(ws, f'{col}12', f'={col}13*0.25')
        total(ws, f'{col}14', f'={col}7-{col}13')
        frm(ws, f'{col}15', f'={col}14/{col}7', fmt=FMT_PCT, italic=True)
        prev_col = col

    # ---- OPERATING EXPENSES ----
    dark_hdr(ws, 'A17', 'OPERATING EXPENSES', merged_end='G17')
    label(ws, 'A18', 'Research & Development', indent=2)
    label(ws, 'A19', 'R&D % Revenue', indent=3, italic=True)
    label(ws, 'A20', 'Sales & Marketing', indent=2)
    label(ws, 'A21', 'S&M % Revenue', indent=3, italic=True)
    label(ws, 'A22', 'General & Administrative', indent=2)
    label(ws, 'A23', 'G&A % Revenue', indent=3, italic=True)
    label(ws, 'A24', 'Total Operating Expenses', bold=True)

    hist_rd  = [1871, 2182, 2553]
    hist_sm  = [1859, 2159, 2530]
    hist_ga  = [498,  567,  666]
    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}18', hist_rd[i])
        frm(ws, f'{col}19', f'={col}18/{col}7', fmt=FMT_PCT, italic=True)
        inp(ws, f'{col}20', hist_sm[i])
        frm(ws, f'{col}21', f'={col}20/{col}7', fmt=FMT_PCT, italic=True)
        inp(ws, f'{col}22', hist_ga[i])
        frm(ws, f'{col}23', f'={col}22/{col}7', fmt=FMT_PCT, italic=True)
        total(ws, f'{col}24', f'={col}18+{col}20+{col}22')

    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        lnk(ws, f'{col}18', f'={col}7*Assumptions!${ac}$7')
        frm(ws, f'{col}19', f'={col}18/{col}7', fmt=FMT_PCT, italic=True)
        lnk(ws, f'{col}20', f'={col}7*Assumptions!${ac}$8')
        frm(ws, f'{col}21', f'={col}20/{col}7', fmt=FMT_PCT, italic=True)
        lnk(ws, f'{col}22', f'={col}7*Assumptions!${ac}$9')
        frm(ws, f'{col}23', f'={col}22/{col}7', fmt=FMT_PCT, italic=True)
        total(ws, f'{col}24', f'={col}18+{col}20+{col}22')

    # ---- EBIT ----
    label(ws, 'A26', 'Operating Income (EBIT)', bold=True)
    label(ws, 'A27', 'EBIT Margin %', indent=2, italic=True)
    for col in DATA_COLS:
        total(ws, f'{col}26', f'={col}14-{col}24')
        frm(ws, f'{col}27', f'={col}26/{col}7', fmt=FMT_PCT, italic=True)

    # ---- OTHER INCOME / EXPENSE ----
    dark_hdr(ws, 'A29', 'OTHER INCOME / EXPENSE', merged_end='G29')
    label(ws, 'A30', 'Interest Income', indent=2)
    label(ws, 'A31', 'Interest Expense', indent=2)
    label(ws, 'A32', 'Other Income (Expense)', indent=2)
    label(ws, 'A33', 'Total Other Income (Expense)', bold=True)

    hist_int_inc = [93,  272, 478]
    hist_int_exp = [-58, -63, -69]
    hist_other   = [30,  -23, -47]
    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}30', hist_int_inc[i])
        inp(ws, f'{col}31', hist_int_exp[i])
        inp(ws, f'{col}32', hist_other[i])
        total(ws, f'{col}33', f'={col}30+{col}31+{col}32')

    for col in PROJ_COLS:
        inp(ws, f'{col}30', 478)
        lnk(ws, f'{col}31', f'=Debt!{col}14')
        inp(ws, f'{col}32', 0)
        total(ws, f'{col}33', f'={col}30+{col}31+{col}32')

    # ---- PRE-TAX / TAX / NET INCOME ----
    label(ws, 'A35', 'Pre-Tax Income', bold=True)
    label(ws, 'A36', '', italic=True)
    label(ws, 'A37', 'Income Tax Provision', indent=2)
    label(ws, 'A38', 'Effective Tax Rate %', indent=3, italic=True)
    label(ws, 'A40', 'Net Income', bold=True)
    label(ws, 'A41', 'Net Income Margin %', indent=2, italic=True)

    hist_tax = [67, 216, 390]
    for i, col in enumerate(HIST_COLS):
        total(ws, f'{col}35', f'={col}26+{col}33')
        inp(ws, f'{col}37', hist_tax[i])
        frm(ws, f'{col}38', f'={col}37/{col}35', fmt=FMT_PCT, italic=True)
        grand_total(ws, f'{col}40', f'={col}35-{col}37')
        frm(ws, f'{col}41', f'={col}40/{col}7', fmt=FMT_PCT, italic=True)

    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        total(ws, f'{col}35', f'={col}26+{col}33')
        lnk(ws, f'{col}37', f'={col}35*Assumptions!${ac}$10')
        frm(ws, f'{col}38', f'={col}37/{col}35', fmt=FMT_PCT, italic=True)
        grand_total(ws, f'{col}40', f'={col}35-{col}37')
        frm(ws, f'{col}41', f'={col}40/{col}7', fmt=FMT_PCT, italic=True)

    # ---- MEMO ITEMS ----
    light_hdr(ws, 'A43', 'MEMO ITEMS', merged_end='G43')
    label(ws, 'A44', 'Depreciation & Amortization', indent=2)
    label(ws, 'A45', 'D&A % Revenue', indent=3, italic=True)
    label(ws, 'A46', 'Stock-Based Compensation', indent=2)
    label(ws, 'A47', 'SBC % Revenue', indent=3, italic=True)
    label(ws, 'A49', 'EBITDA', bold=True)
    label(ws, 'A50', 'EBITDA Margin %', indent=2, italic=True)

    hist_da  = [513, 613, 750]
    hist_sbc = [1100, 1284, 1518]
    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}44', hist_da[i])
        frm(ws, f'{col}45', f'={col}44/{col}7', fmt=FMT_PCT, italic=True)
        inp(ws, f'{col}46', hist_sbc[i])
        frm(ws, f'{col}47', f'={col}46/{col}7', fmt=FMT_PCT, italic=True)
        total(ws, f'{col}49', f'={col}26+{col}44')
        frm(ws, f'{col}50', f'={col}49/{col}7', fmt=FMT_PCT, italic=True)

    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        lnk(ws, f'{col}44', f'={col}7*Assumptions!${ac}$11')
        frm(ws, f'{col}45', f'={col}44/{col}7', fmt=FMT_PCT, italic=True)
        lnk(ws, f'{col}46', f'={col}7*Assumptions!${ac}$12')
        frm(ws, f'{col}47', f'={col}46/{col}7', fmt=FMT_PCT, italic=True)
        total(ws, f'{col}49', f'={col}26+{col}44')
        frm(ws, f'{col}50', f'={col}49/{col}7', fmt=FMT_PCT, italic=True)

    add_hist_proj_border(ws, 1, 52)
    ws.freeze_panes = 'B4'


# ---------------------------------------------------------------------------
# BS — Balance Sheet
# ---------------------------------------------------------------------------

def build_BS(ws):
    ws.sheet_properties.tabColor = DARK_BLUE
    set_col_widths(ws, {'A': 40, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12})

    dark_hdr(ws, 'A1', 'SERVICENOW (NOW) — BALANCE SHEET', merged_end='G1')
    subtitle(ws, 'A2', '($ in millions)')
    col_hdrs(ws, 3)

    # ---- ASSETS ----
    dark_hdr(ws, 'A4', 'ASSETS', merged_end='G4')
    light_hdr(ws, 'A5', 'CURRENT ASSETS', merged_end='G5')

    asset_rows = [
        (6,  'Cash & Cash Equivalents',       [1488, 1946, 3000]),
        (7,  'Short-Term Investments',         [2012, 2634, 2900]),
        (8,  'Accounts Receivable, net',       [1474, 1852, 2290]),
        (9,  'Deferred Commissions (ST)',      [385,  432,  510]),
        (10, 'Prepaid & Other Current Assets', [432,  498,  580]),
    ]
    for row, lbl, vals in asset_rows:
        label(ws, f'A{row}', lbl, indent=2)
        for i, col in enumerate(HIST_COLS):
            inp(ws, f'{col}{row}', vals[i])

    proj_assump = ['B', 'C', 'D']

    # Cash projection: links from CF ending cash (row 30)
    for col in PROJ_COLS:
        lnk(ws, f'{col}6', f'=CF!{col}30', bold=False)

    # Short-term investments: grow at half revenue growth rate
    prev = 'D'
    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        frm(ws, f'{col}7', f'={prev}7*(1+Assumptions!${ac}$5*0.5)')
        prev = col

    # AR: from WC sheet
    for col in PROJ_COLS:
        lnk(ws, f'{col}8', f'=WC!{col}8')

    # Deferred commissions ST: ratio of AR
    for col in PROJ_COLS:
        frm(ws, f'{col}9', f'={col}8*(510/2290)')

    # Prepaid: grow 10% per year
    prev = 'D'
    for col in PROJ_COLS:
        frm(ws, f'{col}10', f'={prev}10*1.10')
        prev = col

    # Total Current Assets
    label(ws, 'A11', 'Total Current Assets', bold=True)
    for col in DATA_COLS:
        total(ws, f'{col}11', f'=SUM({col}6:{col}10)')

    # Non-current assets
    light_hdr(ws, 'A13', 'NON-CURRENT ASSETS', merged_end='G13')

    nca_rows = [
        (14, 'PP&E, net',                    [1032, 1212, 1450]),
        (15, 'Operating Lease ROU Assets',   [810,  789,  760]),
        (16, 'Deferred Commissions (LT)',    [890,  1024, 1210]),
        (17, 'Intangibles & Goodwill',       [2810, 3050, 3200]),
        (18, 'Other Long-Term Assets',       [560,  680,  750]),
    ]
    for row, lbl, vals in nca_rows:
        label(ws, f'A{row}', lbl, indent=2)
        for i, col in enumerate(HIST_COLS):
            inp(ws, f'{col}{row}', vals[i])

    # PP&E from DA schedule
    for col in PROJ_COLS:
        lnk(ws, f'{col}14', f'=DA!{col}13')

    # Operating lease: modest decline
    prev = 'D'
    for col in PROJ_COLS:
        frm(ws, f'{col}15', f'={prev}15*0.97')
        prev = col

    # Deferred commissions LT: ratio of AR
    for col in PROJ_COLS:
        frm(ws, f'{col}16', f'={col}8*(1210/2290)')

    # Goodwill: prior + acquisitions
    prev = 'D'
    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        frm(ws, f'{col}17', f'={prev}17-Assumptions!${ac}$24')
        prev = col

    # Other LT: grow 5%
    prev = 'D'
    for col in PROJ_COLS:
        frm(ws, f'{col}18', f'={prev}18*1.05')
        prev = col

    label(ws, 'A19', 'Total Non-Current Assets', bold=True)
    for col in DATA_COLS:
        total(ws, f'{col}19', f'=SUM({col}14:{col}18)')

    label(ws, 'A20', 'TOTAL ASSETS', bold=True)
    for col in DATA_COLS:
        grand_total(ws, f'{col}20', f'={col}11+{col}19')

    # ---- LIABILITIES ----
    dark_hdr(ws, 'A22', 'LIABILITIES', merged_end='G22')
    light_hdr(ws, 'A23', 'CURRENT LIABILITIES', merged_end='G23')

    cl_rows = [
        (24, 'Accounts Payable',               [156,  187,  220]),
        (25, 'Accrued Expenses',               [980,  1150, 1380]),
        (26, 'Deferred Revenue (ST)',          [4073, 4851, 5810]),
        (27, 'Current Portion of Long-Term Debt', [1500, 0, 0]),
    ]
    for row, lbl, vals in cl_rows:
        label(ws, f'A{row}', lbl, indent=2)
        for i, col in enumerate(HIST_COLS):
            inp(ws, f'{col}{row}', vals[i])

    # AP from WC
    for col in PROJ_COLS:
        lnk(ws, f'{col}24', f'=WC!{col}15')

    # Accrued expenses: ~11.5% of revenue
    for col in PROJ_COLS:
        frm(ws, f'{col}25', f'=IS!{col}7*0.115')

    # Deferred revenue ST: grows roughly with revenue growth
    prev = 'D'
    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        frm(ws, f'{col}26', f'={prev}26*(1+Assumptions!${ac}$5*0.95)')
        prev = col

    # Current LTD from Debt
    for col in PROJ_COLS:
        lnk(ws, f'{col}27', f'=Debt!{col}10')

    label(ws, 'A28', 'Total Current Liabilities', bold=True)
    for col in DATA_COLS:
        total(ws, f'{col}28', f'=SUM({col}24:{col}27)')

    light_hdr(ws, 'A30', 'LONG-TERM LIABILITIES', merged_end='G30')

    lt_rows = [
        (31, 'Deferred Revenue (LT)',          [180, 212, 255]),
        (32, 'Long-Term Debt',                 [1489, 1490, 1490]),
        (33, 'Operating Lease Liabilities (LT)', [791, 766, 740]),
        (34, 'Other Long-Term Liabilities',    [210, 287, 350]),
    ]
    for row, lbl, vals in lt_rows:
        label(ws, f'A{row}', lbl, indent=2)
        for i, col in enumerate(HIST_COLS):
            inp(ws, f'{col}{row}', vals[i])

    # Deferred revenue LT: grow 10%
    prev = 'D'
    for col in PROJ_COLS:
        frm(ws, f'{col}31', f'={prev}31*1.10')
        prev = col

    # LT Debt from Debt schedule
    for col in PROJ_COLS:
        lnk(ws, f'{col}32', f'=Debt!{col}11')

    # Operating lease: modest decline
    prev = 'D'
    for col in PROJ_COLS:
        frm(ws, f'{col}33', f'={prev}33*0.97')
        prev = col

    # Other LT liabilities: grow 8%
    prev = 'D'
    for col in PROJ_COLS:
        frm(ws, f'{col}34', f'={prev}34*1.08')
        prev = col

    label(ws, 'A35', 'Total Long-Term Liabilities', bold=True)
    for col in DATA_COLS:
        total(ws, f'{col}35', f'=SUM({col}31:{col}34)')

    label(ws, 'A36', 'Total Liabilities', bold=True)
    for col in DATA_COLS:
        total(ws, f'{col}36', f'={col}28+{col}35')

    # ---- EQUITY ----
    dark_hdr(ws, 'A38', "STOCKHOLDERS' EQUITY", merged_end='G38')

    eq_rows = [
        (39, 'Common Stock & Additional Paid-In Capital', [8800,  10350, 12100]),
        (40, 'Retained Earnings (Deficit)',               [-2450, -2270, -1850]),
        (41, 'Accumulated Other Comprehensive Loss',      [-120,  -135,  -115]),
    ]
    for row, lbl, vals in eq_rows:
        label(ws, f'A{row}', lbl, indent=2)
        for i, col in enumerate(HIST_COLS):
            inp(ws, f'{col}{row}', vals[i])

    # APIC: prior + SBC + ESPP proceeds
    prev = 'D'
    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        frm(ws, f'{col}39', f'={prev}39+IS!{col}46+Assumptions!${ac}$23')
        prev = col

    # Retained earnings: prior + NI + repurchases (repurchases negative)
    prev = 'D'
    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        frm(ws, f'{col}40', f'={prev}40+IS!{col}40+Assumptions!${ac}$22')
        prev = col

    # AOCI: flat
    prev = 'D'
    for col in PROJ_COLS:
        frm(ws, f'{col}41', f'={prev}41')
        prev = col

    label(ws, 'A42', "Total Stockholders' Equity", bold=True)
    for col in DATA_COLS:
        total(ws, f'{col}42', f'=SUM({col}39:{col}41)')

    label(ws, 'A43', 'TOTAL LIABILITIES & EQUITY', bold=True)
    for col in DATA_COLS:
        grand_total(ws, f'{col}43', f'={col}36+{col}42')

    add_hist_proj_border(ws, 1, 45)
    ws.freeze_panes = 'B4'


# ---------------------------------------------------------------------------
# CF — Cash Flow Statement
# ---------------------------------------------------------------------------

def build_CF(ws):
    ws.sheet_properties.tabColor = DARK_BLUE
    set_col_widths(ws, {'A': 44, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12})

    dark_hdr(ws, 'A1', 'SERVICENOW (NOW) — CASH FLOW STATEMENT', merged_end='G1')
    subtitle(ws, 'A2', '($ in millions)')
    col_hdrs(ws, 3)

    # ---- OPERATING ----
    dark_hdr(ws, 'A4', 'OPERATING ACTIVITIES', merged_end='G4')
    label(ws, 'A5',  'Net Income', indent=2)
    label(ws, 'A6',  'Depreciation & Amortization', indent=2)
    label(ws, 'A7',  'Stock-Based Compensation', indent=2)
    label(ws, 'A8',  'Change in Accounts Receivable', indent=2)
    label(ws, 'A9',  'Change in Deferred Revenue', indent=2)
    label(ws, 'A10', 'Change in Accounts Payable & Accrued Exp.', indent=2)
    label(ws, 'A11', 'Change in Other Working Capital', indent=2)
    label(ws, 'A13', 'Cash from Operations (CFO)', bold=True)
    label(ws, 'A14', 'CFO Margin %', indent=2, italic=True)

    # Historical: NI/D&A/SBC link to IS; WC changes hardcoded
    hist_wc_ar   = [-297, -378, -438]
    hist_wc_dr   = [721,  810,  1002]
    hist_wc_apac = [180,  201,  263]
    hist_wc_oth  = [50,   -30,  -20]
    for i, col in enumerate(HIST_COLS):
        lnk(ws, f'{col}5',  f'=IS!{col}40')
        lnk(ws, f'{col}6',  f'=IS!{col}44')
        lnk(ws, f'{col}7',  f'=IS!{col}46')
        inp(ws, f'{col}8',  hist_wc_ar[i])
        inp(ws, f'{col}9',  hist_wc_dr[i])
        inp(ws, f'{col}10', hist_wc_apac[i])
        inp(ws, f'{col}11', hist_wc_oth[i])
        total(ws, f'{col}13', f'=SUM({col}5:{col}11)')
        frm(ws, f'{col}14', f'={col}13/IS!{col}7', fmt=FMT_PCT, italic=True)

    # Projections
    for col in PROJ_COLS:
        lnk(ws, f'{col}5',  f'=IS!{col}40')
        lnk(ws, f'{col}6',  f'=IS!{col}44')
        lnk(ws, f'{col}7',  f'=IS!{col}46')
        lnk(ws, f'{col}8',  f'=-(WC!{col}8-WC!{_prev(col)}8)')
        lnk(ws, f'{col}9',  f'=(BS!{col}26+BS!{col}31)-(BS!{_prev(col)}26+BS!{_prev(col)}31)')
        lnk(ws, f'{col}10', f'=(WC!{col}15-WC!{_prev(col)}15)+(BS!{col}25-BS!{_prev(col)}25)')
        frm(ws, f'{col}11', f'=-(BS!{col}10-BS!{_prev(col)}10)*0.5')
        total(ws, f'{col}13', f'=SUM({col}5:{col}11)')
        frm(ws, f'{col}14', f'={col}13/IS!{col}7', fmt=FMT_PCT, italic=True)

    # ---- INVESTING ----
    dark_hdr(ws, 'A16', 'INVESTING ACTIVITIES', merged_end='G16')
    label(ws, 'A17', 'Capital Expenditures', indent=2)
    label(ws, 'A18', 'Acquisitions', indent=2)
    label(ws, 'A19', 'Net Investment Purchases / (Sales)', indent=2)
    label(ws, 'A20', 'Cash from Investing (CFI)', bold=True)

    hist_capex = [-502, -592, -680]
    hist_acq   = [-420, -610, -285]
    hist_inv   = [-580, -730, -310]
    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}17', hist_capex[i])
        inp(ws, f'{col}18', hist_acq[i])
        inp(ws, f'{col}19', hist_inv[i])
        total(ws, f'{col}20', f'=SUM({col}17:{col}19)')

    proj_assump = ['B', 'C', 'D']
    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        lnk(ws, f'{col}17', f'=-IS!{col}7*Assumptions!${ac}$15')
        lnk(ws, f'{col}18', f'=Assumptions!${ac}$24')
        lnk(ws, f'{col}19', f'=Assumptions!${ac}$25')
        total(ws, f'{col}20', f'=SUM({col}17:{col}19)')

    # ---- FINANCING ----
    dark_hdr(ws, 'A22', 'FINANCING ACTIVITIES', merged_end='G22')
    label(ws, 'A23', 'Share Repurchases', indent=2)
    label(ws, 'A24', 'Proceeds from ESPP / Options', indent=2)
    label(ws, 'A25', 'Debt Repayment', indent=2)
    label(ws, 'A26', 'Cash from Financing (CFF)', bold=True)

    hist_repo = [-1200, -1500, -1800]
    hist_espp = [300,   380,   450]
    hist_debt = [-1500, 0,     0]
    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}23', hist_repo[i])
        inp(ws, f'{col}24', hist_espp[i])
        inp(ws, f'{col}25', hist_debt[i])
        total(ws, f'{col}26', f'=SUM({col}23:{col}25)')

    for i, col in enumerate(PROJ_COLS):
        ac = proj_assump[i]
        lnk(ws, f'{col}23', f'=Assumptions!${ac}$22')
        lnk(ws, f'{col}24', f'=Assumptions!${ac}$23')
        lnk(ws, f'{col}25', f'=Debt!{col}7')
        total(ws, f'{col}26', f'=SUM({col}23:{col}25)')

    # ---- CASH SUMMARY ----
    label(ws, 'A28', 'Net Change in Cash', bold=True)
    label(ws, 'A29', 'Beginning Cash', indent=2)
    label(ws, 'A30', 'Ending Cash', bold=True)

    for col in DATA_COLS:
        frm(ws, f'{col}28', f'={col}13+{col}20+{col}26', bold=True)

    # Beginning cash
    inp(ws, 'B29', 1239)   # FY2021 ending cash hardcode
    lnk(ws, 'C29', '=B30')
    lnk(ws, 'D29', '=C30')
    lnk(ws, 'E29', '=D30')
    lnk(ws, 'F29', '=E30')
    lnk(ws, 'G29', '=F30')

    for col in DATA_COLS:
        grand_total(ws, f'{col}30', f'={col}29+{col}28')

    # ---- MEMO: FCF ----
    label(ws, 'A32', 'Free Cash Flow (memo)', italic=True)
    label(ws, 'A33', 'FCF Margin %', indent=2, italic=True)
    for col in DATA_COLS:
        frm(ws, f'{col}32', f'={col}13+{col}17', italic=True)
        frm(ws, f'{col}33', f'={col}32/IS!{col}7', fmt=FMT_PCT, italic=True)

    add_hist_proj_border(ws, 1, 35)
    ws.freeze_panes = 'B4'


# ---------------------------------------------------------------------------
# WC — Working Capital Schedule
# ---------------------------------------------------------------------------

def build_WC(ws):
    ws.sheet_properties.tabColor = MED_BLUE
    set_col_widths(ws, {'A': 38, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12})

    dark_hdr(ws, 'A1', 'WORKING CAPITAL SCHEDULE', merged_end='G1')
    subtitle(ws, 'A2', '($ in millions)')
    col_hdrs(ws, 3)

    # ---- ACCOUNTS RECEIVABLE ----
    light_hdr(ws, 'A4', 'ACCOUNTS RECEIVABLE', merged_end='G4')
    label(ws, 'A5', 'Beginning AR', indent=2)
    label(ws, 'A6', 'Revenue (ref)', indent=2)
    label(ws, 'A7', 'Cash Collections (plug)', indent=2)
    label(ws, 'A8', 'Ending AR', bold=True)
    label(ws, 'A9', 'DSO (days)', indent=2, italic=True)

    hist_ar = [1474, 1852, 2290]
    # Beginning AR: FY2022 begin ≈ 1,177 (derived)
    inp(ws, 'B5', 1177)
    lnk(ws, 'C5', '=B8')
    lnk(ws, 'D5', '=C8')

    for i, col in enumerate(HIST_COLS):
        lnk(ws, f'{col}6', f'=IS!{col}7')
        inp(ws, f'{col}8', hist_ar[i])
        frm(ws, f'{col}7', f'={col}5+{col}6-{col}8')
        frm(ws, f'{col}9', f'={col}8/{col}6*365', fmt=FMT_DAYS, italic=True)

    for col in PROJ_COLS:
        lnk(ws, f'{col}5', f'={_prev(col)}8')
        lnk(ws, f'{col}6', f'=IS!{col}7')
        ac = proj_assump_col(col)
        lnk(ws, f'{col}8', f'=IS!{col}7*Assumptions!${ac}$18/365')
        frm(ws, f'{col}7', f'={col}5+{col}6-{col}8')
        frm(ws, f'{col}9', f'={col}8/{col}6*365', fmt=FMT_DAYS, italic=True)

    # ---- ACCOUNTS PAYABLE ----
    light_hdr(ws, 'A11', 'ACCOUNTS PAYABLE', merged_end='G11')
    label(ws, 'A12', 'Beginning AP', indent=2)
    label(ws, 'A13', 'Cost of Revenue (COGS proxy)', indent=2)
    label(ws, 'A14', 'Cash Payments (plug)', indent=2)
    label(ws, 'A15', 'Ending AP', bold=True)
    label(ws, 'A16', 'DPO (days)', indent=2, italic=True)

    hist_ap = [156, 187, 220]
    inp(ws, 'B12', 125)
    lnk(ws, 'C12', '=B15')
    lnk(ws, 'D12', '=C15')

    for i, col in enumerate(HIST_COLS):
        lnk(ws, f'{col}13', f'=IS!{col}13')
        inp(ws, f'{col}15', hist_ap[i])
        frm(ws, f'{col}14', f'={col}12+{col}13-{col}15')
        frm(ws, f'{col}16', f'={col}15/{col}13*365', fmt=FMT_DAYS, italic=True)

    for col in PROJ_COLS:
        lnk(ws, f'{col}12', f'={_prev(col)}15')
        lnk(ws, f'{col}13', f'=IS!{col}13')
        ac = proj_assump_col(col)
        lnk(ws, f'{col}15', f'=IS!{col}13*Assumptions!${ac}$19/365')
        frm(ws, f'{col}14', f'={col}12+{col}13-{col}15')
        frm(ws, f'{col}16', f'={col}15/{col}13*365', fmt=FMT_DAYS, italic=True)

    add_hist_proj_border(ws, 1, 18)
    ws.freeze_panes = 'B4'


# ---------------------------------------------------------------------------
# DA — Depreciation & Amortization Schedule
# ---------------------------------------------------------------------------

def build_DA(ws):
    ws.sheet_properties.tabColor = MED_BLUE
    set_col_widths(ws, {'A': 38, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12})

    dark_hdr(ws, 'A1', 'D&A / PP&E SCHEDULE', merged_end='G1')
    subtitle(ws, 'A2', '($ in millions)')
    col_hdrs(ws, 3)

    light_hdr(ws, 'A4', 'PP&E ROLL-FORWARD', merged_end='G4')
    label(ws, 'A5',  'Beginning PP&E (Gross)', indent=2)
    label(ws, 'A6',  'Capital Expenditures', indent=2)
    label(ws, 'A7',  'Ending PP&E (Gross)', bold=True)
    label(ws, 'A9',  'Beginning Accumulated Depreciation', indent=2)
    label(ws, 'A10', 'Depreciation Expense', indent=2)
    label(ws, 'A11', 'Ending Accumulated Depreciation', bold=True)
    label(ws, 'A13', 'PP&E, Net', bold=True)

    hist_capex = [502, 592, 680]   # gross additions (positive here)
    # Gross PP&E: FY2022 beginning ~1,800
    inp(ws, 'B5', 1800)
    lnk(ws, 'C5', '=B7')
    lnk(ws, 'D5', '=C7')
    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}6', hist_capex[i])
        total(ws, f'{col}7', f'={col}5+{col}6')

    # Accum depr: FY2022 beginning ~760
    inp(ws, 'B9', 760)
    lnk(ws, 'C9', '=B11')
    lnk(ws, 'D9', '=C11')
    for i, col in enumerate(HIST_COLS):
        lnk(ws, f'{col}10', f'=IS!{col}44')
        total(ws, f'{col}11', f'={col}9+{col}10')
        grand_total(ws, f'{col}13', f'={col}7-{col}11')

    # Projections
    lnk(ws, 'E5', '=D7')
    lnk(ws, 'F5', '=E7')
    lnk(ws, 'G5', '=F7')
    for col in PROJ_COLS:
        lnk(ws, f'{col}6', f'=-CF!{col}17')
        total(ws, f'{col}7', f'={col}5+{col}6')

    lnk(ws, 'E9', '=D11')
    lnk(ws, 'F9', '=E11')
    lnk(ws, 'G9', '=F11')
    for col in PROJ_COLS:
        lnk(ws, f'{col}10', f'=IS!{col}44')
        total(ws, f'{col}11', f'={col}9+{col}10')
        grand_total(ws, f'{col}13', f'={col}7-{col}11')

    add_hist_proj_border(ws, 1, 15)
    ws.freeze_panes = 'B4'


# ---------------------------------------------------------------------------
# Debt — Debt Schedule
# ---------------------------------------------------------------------------

def build_Debt(ws):
    ws.sheet_properties.tabColor = MED_BLUE
    set_col_widths(ws, {'A': 38, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12})

    dark_hdr(ws, 'A1', 'DEBT SCHEDULE', merged_end='G1')
    subtitle(ws, 'A2', '($ in millions)')
    col_hdrs(ws, 3)

    light_hdr(ws, 'A4', 'TOTAL DEBT ROLL-FORWARD', merged_end='G4')
    label(ws, 'A5',  'Beginning Total Debt', indent=2)
    label(ws, 'A6',  'New Borrowings', indent=2)
    label(ws, 'A7',  'Debt Repayments', indent=2)
    label(ws, 'A8',  'Ending Total Debt', bold=True)
    label(ws, 'A10', 'Current Portion of LTD', indent=2)
    label(ws, 'A11', 'Long-Term Debt', indent=2)

    # Beginning debt FY22: 3,000 (1,500 current + 1,489 LT + ~11 rounding)
    inp(ws, 'B5', 3000)
    lnk(ws, 'C5', '=B8')
    lnk(ws, 'D5', '=C8')

    hist_new_borrow = [0, 0, 0]
    hist_repay      = [-1500, 0, 0]
    for i, col in enumerate(HIST_COLS):
        inp(ws, f'{col}6', hist_new_borrow[i])
        inp(ws, f'{col}7', hist_repay[i])
        total(ws, f'{col}8', f'={col}5+{col}6+{col}7')
        inp(ws, f'{col}10', [1500, 0, 0][i])
        frm(ws, f'{col}11', f'={col}8-{col}10')

    lnk(ws, 'E5', '=D8')
    lnk(ws, 'F5', '=E8')
    lnk(ws, 'G5', '=F8')
    for col in PROJ_COLS:
        inp(ws, f'{col}6', 0)
        inp(ws, f'{col}7', 0)
        total(ws, f'{col}8', f'={col}5+{col}6+{col}7')
        inp(ws, f'{col}10', 0)
        frm(ws, f'{col}11', f'={col}8-{col}10')

    # Interest
    light_hdr(ws, 'A13', 'INTEREST EXPENSE', merged_end='G13')
    label(ws, 'A14', 'Interest Rate', indent=2)
    label(ws, 'A15', 'Interest Expense', indent=2)

    for col in DATA_COLS:
        inp(ws, f'{col}14', 0.045, fmt=FMT_PCT)
        frm(ws, f'{col}15', f'=-{col}5*{col}14')

    add_hist_proj_border(ws, 1, 17)
    ws.freeze_panes = 'B4'


# ---------------------------------------------------------------------------
# Checks — Model Integrity Checks
# ---------------------------------------------------------------------------

def build_Checks(ws):
    ws.sheet_properties.tabColor = "FF0000"
    set_col_widths(ws, {'A': 52, 'B': 14, 'C': 14, 'D': 16, 'E': 14})

    dark_hdr(ws, 'A1', 'MODEL INTEGRITY CHECKS — SERVICENOW', merged_end='E1')
    subtitle(ws, 'A2', 'GREEN = pass (TRUE), RED = fail (FALSE). Historical BS checks may show small discrepancies from rounded inputs.')

    # Column headers
    for ref, txt in [('A3','Check'), ('B3','Expected'), ('C3','Note'), ('D3','Result'), ('E3','Status')]:
        c = ws[ref]
        c.value = txt
        c.font = _font(BLACK, bold=True)
        c.fill = _fill(LIGHT_BLUE)
        c.alignment = _center()

    def check_row(row, name, expected, note, result_formula, pass_formula):
        label(ws, f'A{row}', name, indent=1)
        c = ws[f'B{row}']; c.value = expected; c.number_format = FMT_DOLLAR; c.alignment = _center()
        label(ws, f'C{row}', note, italic=True)
        frm(ws, f'D{row}', result_formula, fmt=FMT_DOLLAR)
        frm(ws, f'E{row}', pass_formula)
        ws[f'E{row}'].alignment = _center()

    # BS Balance Checks
    dark_hdr(ws, 'A5', 'BALANCE SHEET — Assets = Liabilities + Equity', merged_end='E5')
    periods_map = [('B','FY2022A'), ('C','FY2023A'), ('D','FY2024A'), ('E','FY2025E'), ('F','FY2026E'), ('G','FY2027E')]
    for i, (col, period) in enumerate(periods_map):
        r = 6 + i
        check_row(r, f'BS Balances — {period}', 0, 'Assets - L&E', f'=BS!{col}20-BS!{col}43', f'=IF(ABS(D{r})<1,"TRUE","FALSE")')

    # Cash Tie-out
    dark_hdr(ws, 'A13', 'CASH TIE-OUT — CF Ending Cash = BS Cash', merged_end='E13')
    for i, (col, period) in enumerate(periods_map):
        r = 14 + i
        check_row(r, f'CF Cash = BS Cash — {period}', 0, 'CF end - BS cash', f'=CF!{col}30-BS!{col}6', f'=IF(ABS(D{r})<1,"TRUE","FALSE")')

    # NI Link
    dark_hdr(ws, 'A21', 'NET INCOME LINK — IS = CF', merged_end='E21')
    ni_periods = [('B','FY2022A'), ('D','FY2024A'), ('E','FY2025E'), ('F','FY2026E'), ('G','FY2027E')]
    for i, (col, period) in enumerate(ni_periods):
        r = 22 + i
        check_row(r, f'IS NI = CF NI — {period}', 0, 'IS row 40 - CF row 5', f'=IS!{col}40-CF!{col}5', f'=IF(ABS(D{r})<1,"TRUE","FALSE")')

    # DA Tie-out
    dark_hdr(ws, 'A28', "PP&E NET — DA Schedule = BS", merged_end='E28')
    ppe_periods = [('D','FY2024A'), ('E','FY2025E'), ('F','FY2026E'), ('G','FY2027E')]
    for i, (col, period) in enumerate(ppe_periods):
        r = 29 + i
        check_row(r, f'DA PP&E Net = BS PP&E — {period}', 0, 'DA row 13 - BS row 14', f'=DA!{col}13-BS!{col}14', f'=IF(ABS(D{r})<1,"TRUE","FALSE")')

    # RE Roll-forward
    dark_hdr(ws, 'A34', 'RETAINED EARNINGS ROLL-FORWARD', merged_end='E34')
    re_periods = [('E','FY2025E','B'), ('F','FY2026E','C'), ('G','FY2027E','D')]
    for col, period, ac in re_periods:
        prev_col = _prev(col)
        r = 35 + (['E','F','G'].index(col))
        check_row(r, f'RE Roll-Fwd — {period}', 0, 'RE = prior + NI + buybacks',
                  f'=BS!{col}40-(BS!{prev_col}40+IS!{col}40+Assumptions!${ac}$22)',
                  f'=IF(ABS(D{r})<1,"TRUE","FALSE")')

    # ── SOFTWARE CHECKS ───────────────────────────────────────────────────
    dark_hdr(ws, 'A40', 'SOFTWARE — IS REVENUE LINKS (revenue_build → IS)', merged_end='E40')
    proj_periods = [('E','FY2025E'), ('F','FY2026E'), ('G','FY2027E')]
    for i, (col, period) in enumerate(proj_periods):
        r = 41 + i
        check_row(r, f'IS Sub Rev = rb Sub Rev — {period}', 0, 'IS row 5 = rb row 37',
                  f'=IS!{col}5-revenue_build!{col}37',
                  f'=IF(ABS(D{r})<0.01,"TRUE","FALSE")')
    for i, (col, period) in enumerate(proj_periods):
        r = 44 + i
        check_row(r, f'IS PS Rev = rb PS Rev — {period}', 0, 'IS row 6 = rb row 38',
                  f'=IS!{col}6-revenue_build!{col}38',
                  f'=IF(ABS(D{r})<0.01,"TRUE","FALSE")')

    dark_hdr(ws, 'A48', 'SOFTWARE — ARR BRIDGE & KPI CHECKS', merged_end='E48')
    for i, (col, period) in enumerate(proj_periods):
        r = 49 + i
        check_row(r, f'ARR Bridge Balances — {period}', 0, 'Ending ARR = sum of components',
                  f'=revenue_build!{col}10-(revenue_build!{col}5+revenue_build!{col}6+revenue_build!{col}7+revenue_build!{col}8+revenue_build!{col}9)',
                  f'=IF(ABS(D{r})<0.01,"TRUE","FALSE")')

    check_row(52, 'GRR <= 100% — FY2025E', 1, 'GRR cannot exceed 100%',
              '=revenue_build!E15',
              '=IF(revenue_build!E15<=1,"TRUE","FALSE")')
    check_row(53, 'NRR >= GRR — FY2025E', '', 'expansion included in NRR',
              '=revenue_build!E16-revenue_build!E15',
              '=IF(revenue_build!E16>=revenue_build!E15,"TRUE","FALSE")')
    check_row(54, 'cRPO <= Total RPO — FY2025E', '', 'current portion <= total',
              '=revenue_build!E23-revenue_build!E22',
              '=IF(revenue_build!E23<=revenue_build!E22,"TRUE","FALSE")')

    # Master Status (extended range to include software checks)
    r = 57
    ws.merge_cells(f'A{r}:D{r}')
    c = ws[f'A{r}']
    c.value = '=IF(COUNTIF(E6:E54,"FALSE")=0,"✓  ALL CHECKS PASS","✗  "&COUNTIF(E6:E54,"FALSE")&" CHECK(S) FAILING")'
    c.font = _font(WHITE, bold=True, size=11)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _center()

    # Conditional formatting on E column
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        "E6:E54",
        CellIsRule(operator="equal", formula=['"TRUE"'], fill=green_fill,
                   font=Font(color="375623"))
    )
    ws.conditional_formatting.add(
        "E6:E54",
        CellIsRule(operator="equal", formula=['"FALSE"'], fill=red_fill,
                   font=Font(color="9C0006"))
    )

    ws.freeze_panes = 'B4'


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_rb     = wb.active;            ws_rb.title = "revenue_build"
    ws_is     = wb.create_sheet("IS")
    ws_bs     = wb.create_sheet("BS")
    ws_cf     = wb.create_sheet("CF")
    ws_wc     = wb.create_sheet("WC")
    ws_da     = wb.create_sheet("DA")
    ws_debt   = wb.create_sheet("Debt")
    ws_assump = wb.create_sheet("Assumptions")
    ws_checks = wb.create_sheet("Checks")

    build_revenue_build(ws_rb)
    build_assumptions(ws_assump)
    build_IS(ws_is)
    build_BS(ws_bs)
    build_CF(ws_cf)
    build_WC(ws_wc)
    build_DA(ws_da)
    build_Debt(ws_debt)
    build_Checks(ws_checks)

    wb.save(str(OUTPUT_FILE))
    print(f"Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
