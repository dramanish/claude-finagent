import importlib.util
import json
import os
import threading
import subprocess
import sys
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from types import SimpleNamespace

import jwt
import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]


def load_module(name: str, relative_path: str):
    path = ROOT / relative_path
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


def test_validate_loads_json_and_yaml_and_reports_validation_errors(tmp_path, capsys):
    mod = load_module("validate_script", "scripts/validate.py")
    instance = tmp_path / "out.json"
    schema = tmp_path / "schema.yaml"
    instance.write_text(json.dumps({"count": "bad"}))
    schema.write_text("type: object\nproperties:\n  count:\n    type: integer\nrequired: [count]\n")

    sys.argv = ["validate.py", str(instance), str(schema)]
    assert mod.main() == 1
    assert "INVALID:" in capsys.readouterr().err

    instance.write_text(json.dumps({"count": 3}))
    assert mod.main() == 0
    assert capsys.readouterr().out.strip() == "OK"

    sys.argv = ["validate.py"]
    assert mod.main() == 2


def test_orchestrate_extract_handoff_rejects_bad_inputs_and_accepts_valid_payload():
    mod = load_module("orchestrate_script", "scripts/orchestrate.py")
    assert mod.extract_handoff("no handoff") is None
    assert mod.extract_handoff('{"type":"handoff_request", broken') is None
    assert mod.extract_handoff('{"type":"handoff_request","target_agent":"bad","payload":{"event":"x"}}') is None
    assert mod.extract_handoff('{"type":"handoff_request","target_agent":"pitch-agent","payload":{"event":3}}') is None

    handoff = mod.extract_handoff(
        'prefix {"type":"handoff_request","target_agent":"pitch-agent",'
        '"payload":{"event":"review this","context_ref":"deal/1#slide-2"}} suffix'
    )
    assert handoff == {
        "target_agent": "pitch-agent",
        "payload": {"event": "review this", "context_ref": "deal/1#slide-2"},
    }


def test_orchestrate_run_steers_only_allowed_handoffs(monkeypatch):
    mod = load_module("orchestrate_run_script", "scripts/orchestrate.py")
    events = [
        SimpleNamespace(type="message_delta", text="plain"),
        SimpleNamespace(
            type="message_delta",
            text='{"type":"handoff_request","target_agent":"pitch-agent","payload":{"event":"go"}}',
        ),
        SimpleNamespace(type="other", text='{"type":"handoff_request","target_agent":"pitch-agent","payload":{"event":"skip"}}'),
    ]
    steered = []

    class Stream:
        def __enter__(self):
            return iter(events)

        def __exit__(self, exc_type, exc, tb):
            return False

    class Sessions:
        def stream(self, session_id):
            assert session_id == "source"
            return Stream()

        def steer(self, agent_id, input):
            steered.append((agent_id, input))

    monkeypatch.setattr(mod.anthropic, "Anthropic", lambda: SimpleNamespace(beta=SimpleNamespace(agents=SimpleNamespace(sessions=Sessions()))))
    mod.run("source", {"pitch-agent": "agent-1"})
    assert steered == [("agent-1", "go")]
    steered.clear()
    mod.run("source", {})
    assert steered == []


def test_sync_agent_skills_copies_sources_and_reports_missing(tmp_path):
    mod = load_module("sync_agent_skills", "scripts/sync-agent-skills.py")
    src = tmp_path / "plugins" / "vertical-plugins" / "finance" / "skills" / "shared"
    dst = tmp_path / "plugins" / "agent-plugins" / "agent" / "skills" / "shared"
    missing = tmp_path / "plugins" / "agent-plugins" / "agent" / "skills" / "missing"
    src.mkdir(parents=True)
    dst.mkdir(parents=True)
    missing.mkdir(parents=True)
    (src / "SKILL.md").write_text("new")
    (dst / "old.txt").write_text("old")
    (tmp_path / "plugins" / "vertical-plugins" / "finance" / "skills" / "not-a-dir").write_text("x")
    (tmp_path / "plugins" / "agent-plugins" / "agent" / "skills" / "not-a-dir").write_text("x")

    assert mod.sync_agent_skills(tmp_path) == (1, ["plugins/agent-plugins/agent/skills/missing"])
    assert (dst / "SKILL.md").read_text() == "new"
    assert not (dst / "old.txt").exists()
    monkeypatch = pytest.MonkeyPatch()
    try:
        monkeypatch.setattr(mod, "sync_agent_skills", lambda: (0, []))
        assert mod.main() == 0
        monkeypatch.setattr(mod, "sync_agent_skills", lambda: (0, ["missing"]))
        assert mod.main() == 1
    finally:
        monkeypatch.undo()


def test_get_tenant_id_from_domain_and_az_cli(monkeypatch):
    mod = load_module("get_tenant_id", "claude-for-msft-365-install/examples/python-bootstrap/get_tenant_id.py")

    class Response:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return b'{"issuer":"https://login.microsoftonline.com/tenant-123/v2.0"}'

    monkeypatch.setattr(mod.urllib.request, "urlopen", lambda url, timeout: Response())
    assert mod.from_domain("example.com") == "tenant-123"

    monkeypatch.setattr(mod.subprocess, "run", lambda *a, **k: SimpleNamespace(stdout="tenant-abc\n"))
    assert mod.from_az_cli() == "tenant-abc"


def test_bootstrap_config_resolution_and_token_validation(tmp_path, monkeypatch):
    monkeypatch.setenv("TENANT_ID", "dev-tenant")
    monkeypatch.setenv("DEV_JWKS_PATH", str(tmp_path / "dev_jwks.json"))
    monkeypatch.setenv("HOST", "127.0.0.1")
    bootstrap_dir = ROOT / "claude-for-msft-365-install" / "examples" / "python-bootstrap"
    sys.path.insert(0, str(bootstrap_dir))
    for name in ("app", "config", "mint_dev_token"):
        sys.modules.pop(name, None)
    try:
        mint = load_module("mint_dev_token", "claude-for-msft-365-install/examples/python-bootstrap/mint_dev_token.py")
        token = mint.mint_token("bob", ["risk"], str(tmp_path / "dev_private.pem"), str(tmp_path / "dev_jwks.json"))
        app = load_module("app", "claude-for-msft-365-install/examples/python-bootstrap/app.py")

        assert app.parse_app("Claude-Excel/1.0") == "excel"
        assert app.parse_app("unknown") == ""
        assert app.resolve("alice", set(), "")["skills"][0]["name"] == "deal-memo"
        assert app.resolve("bob", {"risk"}, "excel")["mcp_servers"][0]["label"] == "Risk Dashboard"
        assert app.resolve("bob", {"investment-banking"}, "word")["skills"][0]["name"] == "deal-memo"
        assert app.resolve("nobody", set(), "")["skills"][0]["name"] == "compliance-check"

        claims = app.validate(f"Bearer {token}")
        assert claims["oid"] == "bob"
        with pytest.raises(HTTPException):
            app.validate("")
        with pytest.raises(HTTPException):
            app.validate("Bearer broken")
        response = app.bootstrap(f"Bearer {token}", "Claude-Excel/1.0")
        assert response["mcp_servers"][0]["label"] == "Risk Dashboard"
        assert response["bootstrap_expires_at"] > 0
        monkeypatch.setattr(app, "RULES", [{"when": {"app": "word"}, "skills": [], "mcp_servers": []}])
        assert app.resolve("bob", set(), "excel") == {"skills": [], "mcp_servers": []}
        monkeypatch.setattr(app, "DEV_JWKS_PATH", "")
        monkeypatch.setattr(app, "_jwks", SimpleNamespace(get_signing_key_from_jwt=lambda token: SimpleNamespace(key="secret")))
        monkeypatch.setattr(app.jwt, "decode", lambda *args, **kwargs: {"oid": "remote"})
        assert app.validate("Bearer remote-token") == {"oid": "remote"}
    finally:
        sys.path.remove(str(bootstrap_dir))


def test_config_rejects_dev_jwks_on_non_loopback(monkeypatch, tmp_path):
    monkeypatch.setenv("TENANT_ID", "dev-tenant")
    monkeypatch.setenv("DEV_JWKS_PATH", str(tmp_path / "jwks.json"))
    monkeypatch.setenv("HOST", "0.0.0.0")
    with pytest.raises(SystemExit):
        load_module("config_bad_host", "claude-for-msft-365-install/examples/python-bootstrap/config.py")


def test_mint_dev_token_creates_reusable_key_and_jwks(tmp_path, monkeypatch):
    monkeypatch.setenv("TENANT_ID", "dev-tenant")
    mod = load_module("mint_dev_token_reuse", "claude-for-msft-365-install/examples/python-bootstrap/mint_dev_token.py")
    priv = tmp_path / "dev_private.pem"
    jwks = tmp_path / "dev_jwks.json"
    token1 = mod.mint_token("alice", ["g1"], str(priv), str(jwks))
    token2 = mod.mint_token("alice", ["g2"], str(priv), str(jwks))
    assert priv.exists()
    assert json.loads(jwks.read_text())["keys"][0]["kid"] == "dev"
    assert jwt.get_unverified_header(token1)["kid"] == "dev"
    assert jwt.decode(token2, options={"verify_signature": False})["groups"] == ["g2"]
    assert mod.b64u(65537) == "AQAB"
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", ["mint_dev_token.py", "--oid", "cli", "--group", "g"])
    mod.main()


def test_quick_validate_all_frontmatter_paths(tmp_path):
    mod = load_module(
        "quick_validate",
        "plugins/vertical-plugins/financial-analysis/skills/skill-creator/scripts/quick_validate.py",
    )
    skill = tmp_path / "skill"
    skill.mkdir()
    assert mod.validate_skill(skill) == (False, "SKILL.md not found")
    skill_md = skill / "SKILL.md"
    skill_md.write_text("body")
    assert mod.validate_skill(skill)[1] == "No YAML frontmatter found"
    skill_md.write_text("---\nname: x\n")
    assert mod.validate_skill(skill)[1] == "Invalid frontmatter format"
    skill_md.write_text("---\n[]\n---\n")
    assert mod.validate_skill(skill)[1] == "Frontmatter must be a YAML dictionary"
    skill_md.write_text("---\nname: [\n---\n")
    assert "Invalid YAML in frontmatter" in mod.validate_skill(skill)[1]
    skill_md.write_text("---\nname: good\nextra: x\n---\n")
    assert "Unexpected key" in mod.validate_skill(skill)[1]
    skill_md.write_text("---\ndescription: desc\n---\n")
    assert mod.validate_skill(skill)[1] == "Missing 'name' in frontmatter"
    skill_md.write_text("---\nname: good\n---\n")
    assert mod.validate_skill(skill)[1] == "Missing 'description' in frontmatter"
    skill_md.write_text("---\nname: 3\ndescription: desc\n---\n")
    assert mod.validate_skill(skill)[1] == "Name must be a string, got int"
    skill_md.write_text("---\nname: Bad_Name\ndescription: desc\n---\n")
    assert "hyphen-case" in mod.validate_skill(skill)[1]
    skill_md.write_text("---\nname: bad--name\ndescription: desc\n---\n")
    assert "cannot start/end" in mod.validate_skill(skill)[1]
    skill_md.write_text(f"---\nname: {'a' * 65}\ndescription: desc\n---\n")
    assert "too long" in mod.validate_skill(skill)[1]
    skill_md.write_text("---\nname: good\ndescription: 5\n---\n")
    assert mod.validate_skill(skill)[1] == "Description must be a string, got int"
    skill_md.write_text("---\nname: good\ndescription: bad <tag>\n---\n")
    assert "angle brackets" in mod.validate_skill(skill)[1]
    skill_md.write_text(f"---\nname: good\ndescription: {'x' * 1025}\n---\n")
    assert "Description is too long" in mod.validate_skill(skill)[1]
    skill_md.write_text("---\nname: good-skill\ndescription: useful skill\n---\n")
    assert mod.validate_skill(skill) == (True, "Skill is valid!")
    skill_md.write_text("---\nname: ''\ndescription: ''\n---\n")
    assert mod.validate_skill(skill) == (True, "Skill is valid!")


def test_init_and_package_skill(tmp_path, capsys, monkeypatch):
    skill_dir = ROOT / "plugins/vertical-plugins/financial-analysis/skills/skill-creator/scripts"
    sys.path.insert(0, str(skill_dir))
    sys.modules.pop("quick_validate", None)
    try:
        init = load_module("init_skill", "plugins/vertical-plugins/financial-analysis/skills/skill-creator/scripts/init_skill.py")
        package = load_module("package_skill", "plugins/vertical-plugins/financial-analysis/skills/skill-creator/scripts/package_skill.py")
        assert init.title_case_skill_name("new-data-tool") == "New Data Tool"
        created = init.init_skill("new-data-tool", tmp_path)
        assert (created / "SKILL.md").exists()
        assert (created / "scripts" / "example.py").stat().st_mode & 0o111
        assert init.init_skill("new-data-tool", tmp_path) is None
        (created / "SKILL.md").write_text("---\nname: new-data-tool\ndescription: useful skill\n---\n")
        bad_base = tmp_path / "bad"
        monkeypatch.setattr(init.Path, "mkdir", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("mkdir fail")))
        assert init.init_skill("mkdir-fail", bad_base) is None
        monkeypatch.undo()
        skill_dir = ROOT / "plugins/vertical-plugins/financial-analysis/skills/skill-creator/scripts"
        sys.path.insert(0, str(skill_dir))
        init = load_module("init_skill_again", "plugins/vertical-plugins/financial-analysis/skills/skill-creator/scripts/init_skill.py")
        real_write_text = init.Path.write_text
        monkeypatch.setattr(init.Path, "write_text", lambda self, text: (_ for _ in ()).throw(RuntimeError("write fail")) if self.name == "SKILL.md" else real_write_text(self, text))
        assert init.init_skill("skill-write-fail", tmp_path) is None
        monkeypatch.undo()
        init = load_module("init_skill_resource", "plugins/vertical-plugins/financial-analysis/skills/skill-creator/scripts/init_skill.py")
        real_write_text = init.Path.write_text
        monkeypatch.setattr(init.Path, "write_text", lambda self, text: (_ for _ in ()).throw(RuntimeError("resource fail")) if self.name == "example_asset.txt" else real_write_text(self, text))
        assert init.init_skill("resource-write-fail", tmp_path) is None
        monkeypatch.undo()

        out_dir = tmp_path / "dist"
        (created / "nested-dir").mkdir()
        packaged = package.package_skill(created, out_dir)
        assert packaged.exists()
        assert packaged.suffix == ".skill"
        assert package.package_skill(tmp_path / "missing") is None
        not_dir = tmp_path / "file"
        not_dir.write_text("x")
        assert package.package_skill(not_dir) is None
        no_skill = tmp_path / "no-skill"
        no_skill.mkdir()
        assert package.package_skill(no_skill) is None
        monkeypatch.setattr(package, "validate_skill", lambda path: (False, "bad"))
        assert package.package_skill(created) is None
        monkeypatch.setattr(package, "validate_skill", lambda path: (True, "ok"))
        monkeypatch.setattr(package.zipfile, "ZipFile", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("zip fail")))
        assert package.package_skill(created) is None
        monkeypatch.setattr(sys, "argv", ["package_skill.py"])
        with pytest.raises(SystemExit) as exc:
            package.main()
        assert exc.value.code == 1
        monkeypatch.setattr(sys, "argv", ["package_skill.py", str(created), str(out_dir)])
        monkeypatch.setattr(package, "package_skill", lambda *args: out_dir / "ok.skill")
        with pytest.raises(SystemExit) as exc:
            package.main()
        assert exc.value.code == 0
        monkeypatch.setattr(sys, "argv", ["package_skill.py", str(created)])
        with pytest.raises(SystemExit) as exc:
            package.main()
        assert exc.value.code == 0
        monkeypatch.setattr(package, "package_skill", lambda *args: None)
        with pytest.raises(SystemExit) as exc:
            package.main()
        assert exc.value.code == 1
        monkeypatch.setattr(sys, "argv", ["init_skill.py"])
        with pytest.raises(SystemExit) as exc:
            init.main()
        assert exc.value.code == 1
        monkeypatch.setattr(sys, "argv", ["init_skill.py", "cli-skill", "--path", str(tmp_path)])
        with pytest.raises(SystemExit) as exc:
            init.main()
        assert exc.value.code == 0
        monkeypatch.setattr(init, "init_skill", lambda *args: None)
        monkeypatch.setattr(sys, "argv", ["init_skill.py", "bad-skill", "--path", str(tmp_path)])
        with pytest.raises(SystemExit) as exc:
            init.main()
        assert exc.value.code == 1
    finally:
        sys.path.remove(str(skill_dir))
    assert "Created SKILL.md" in capsys.readouterr().out


def test_extract_numbers_normalization_categories_inconsistencies_and_cli(tmp_path, capsys):
    mod = load_module(
        "extract_numbers",
        "plugins/vertical-plugins/financial-analysis/skills/ib-check-deck/scripts/extract_numbers.py",
    )
    assert mod.normalize_number("bad", "M") == 0.0
    assert mod.normalize_number("1.5", "billion") == 1.5e9
    assert mod.detect_category("EV/EBITDA multiple", "x") == "ebitda"
    assert mod.detect_category("EBITDA margin", "%") == "ebitda_margin"
    assert mod.detect_category("trading multiple", "x") == "multiple"
    assert mod.detect_category("profit margin", "%") == "margin"
    assert mod.detect_category("CAGR growth", "%") == "growth"
    assert mod.detect_category("market cap", "USD_M") == "valuation"
    assert mod.detect_category("plain", "%") == "percentage"
    assert mod.detect_category("plain", "x") == "multiple"
    assert mod.detect_category("plain", "none") == "other"
    assert mod.find_inconsistencies([mod.NumberInstance("1", 1, "none", 1, "x", 1, "other")]) == []
    assert mod.find_inconsistencies([mod.NumberInstance("10", 10, "M", 1, "revenue", 1, "revenue")]) == []
    assert mod.find_inconsistencies([
        mod.NumberInstance("10", 10, "M", 1, "revenue", 1, "revenue"),
        mod.NumberInstance("10.1", 10.1, "M", 2, "revenue", 2, "revenue"),
    ]) == []

    content = """# Slide 1
Revenue was $500M and EBITDA margin was 25%. Also $42 standalone and 5 ignored.
<!-- Slide 2 -->
Revenue was $700M and 2024 had no unit.
Revenue was $700M again.
Sales were $0M and sales were $10M.
"""
    numbers = mod.extract_numbers(content)
    values = [n.value for n in numbers]
    assert "$500M" in values
    assert "$700M" in values
    assert "2024" not in values
    inconsistencies = mod.find_inconsistencies(numbers)
    assert inconsistencies[0]["category"] == "revenue"
    assert inconsistencies[0]["severity"] == "high"

    input_file = tmp_path / "deck.md"
    output_file = tmp_path / "numbers.json"
    input_file.write_text(content)
    sys.argv = ["extract_numbers.py", str(input_file), "--output", str(output_file), "--check"]
    mod.main()
    assert json.loads(output_file.read_text())["inconsistencies"]
    assert "POTENTIAL INCONSISTENCIES" in capsys.readouterr().err
    sys.argv = ["extract_numbers.py", str(input_file)]
    mod.main()
    assert '"total_numbers"' in capsys.readouterr().out
    sys.argv = ["extract_numbers.py", str(input_file), "--check"]
    mod.main()
    assert '"inconsistencies"' in capsys.readouterr().out
    clean_file = tmp_path / "clean.md"
    clean_file.write_text("# Slide 1\nRevenue was $500M.\n# Slide 2\nRevenue was $501M.\n")
    sys.argv = ["extract_numbers.py", str(clean_file), "--check"]
    mod.main()
    assert '"inconsistencies": []' in capsys.readouterr().out

    sys.argv = ["extract_numbers.py", str(tmp_path / "missing.md")]
    with pytest.raises(SystemExit):
        mod.main()


def make_dcf_workbook(path: Path, terminal_growth=0.03, wacc=0.10, terminal_value=60, enterprise_value=100):
    wb = Workbook()
    dcf = wb.active
    dcf.title = "DCF"
    dcf["A1"] = "Terminal growth"
    dcf["B1"] = terminal_growth
    dcf["A2"] = "WACC"
    dcf["B2"] = wacc
    dcf["A3"] = "PV terminal value"
    dcf["B3"] = terminal_value
    dcf["A4"] = "Enterprise value"
    dcf["B4"] = enterprise_value
    wb.create_sheet("WACC")["A1"] = "WACC"
    wb["WACC"]["B1"] = wacc
    wb.create_sheet("Sensitivity")
    wb.save(path)


def test_dcf_validator_pass_warnings_errors_and_cli(tmp_path, capsys):
    mod = load_module(
        "validate_dcf",
        "plugins/vertical-plugins/financial-analysis/skills/dcf-model/scripts/validate_dcf.py",
    )
    import builtins
    real_import = builtins.__import__
    def fake_import(name, *args, **kwargs):
        if name == "openpyxl":
            raise ImportError("missing")
        return real_import(name, *args, **kwargs)
    monkeypatch = pytest.MonkeyPatch()
    monkeypatch.setattr(builtins, "__import__", fake_import)
    with pytest.raises(ImportError):
        mod.DCFModelValidator("anything.xlsx")
    monkeypatch.undo()
    good = tmp_path / "good.xlsx"
    make_dcf_workbook(good)
    wb = load_workbook(good)
    wb["DCF"]["C1"] = "=1+1"
    wb.save(good)
    result = mod.validate_dcf_model(str(good))
    assert result["status"] == "PASS"
    assert any("Terminal growth" in item for item in result["info"])

    bad = tmp_path / "bad.xlsx"
    make_dcf_workbook(bad, terminal_growth=0.12, wacc=0.04, terminal_value=90)
    result = mod.validate_dcf_model(str(bad))
    assert result["status"] == "FAIL"
    assert any("CRITICAL" in item for item in result["errors"])
    assert any("outside typical range" in item for item in result["warnings"])
    assert any("over-reliant" in item for item in result["warnings"])
    fake = object.__new__(mod.DCFModelValidator)
    fake.warnings = []
    class ExplodingWorkbook:
        def __getitem__(self, key):
            return SimpleNamespace(iter_rows=lambda **kwargs: (_ for _ in ()).throw(RuntimeError("boom")))
    fake.workbook_values = ExplodingWorkbook()
    mod.DCFModelValidator._check_terminal_growth_vs_wacc(fake)
    assert "Could not validate terminal growth vs WACC: boom" in fake.warnings

    sparse = tmp_path / "sparse.xlsx"
    wb = Workbook()
    wb.active.title = "Only"
    wb.active["A1"] = "#REF!"
    wb.save(sparse)
    result = mod.validate_dcf_model(str(sparse))
    assert result["status"] == "FAIL"
    assert any("Recommended sheet missing" in item for item in result["warnings"])
    assert any("#REF!" in item for item in result["errors"])

    low_tv = tmp_path / "low_tv.xlsx"
    make_dcf_workbook(low_tv, terminal_value=20)
    assert any("too conservative" in item for item in mod.validate_dcf_model(str(low_tv))["warnings"])

    missing_values = tmp_path / "missing_values.xlsx"
    wb = Workbook()
    dcf = wb.active
    dcf.title = "DCF"
    dcf["A1"] = "Terminal growth"
    dcf["A2"] = "WACC"
    dcf["A3"] = "PV terminal value"
    dcf["A4"] = "Enterprise value"
    wb.create_sheet("Sensitivity")
    wb.save(missing_values)
    missing_result = mod.validate_dcf_model(str(missing_values))
    assert "Could not locate terminal growth and WACC values" in missing_result["warnings"]
    assert "Could not locate WACC value" in missing_result["warnings"]
    assert "Could not locate terminal value and enterprise value" in missing_result["warnings"]

    sys.argv = ["validate_dcf.py"]
    with pytest.raises(SystemExit) as exc:
        mod.main()
    assert exc.value.code == 1

    sys.argv = ["validate_dcf.py", str(good), str(tmp_path / "result.json")]
    with pytest.raises(SystemExit) as exc:
        mod.main()
    assert exc.value.code == 0
    assert (tmp_path / "result.json").exists()
    sys.argv = ["validate_dcf.py", str(good)]
    with pytest.raises(SystemExit) as exc:
        mod.main()
    assert exc.value.code == 0

    sys.argv = ["validate_dcf.py", str(tmp_path / "missing.xlsx")]
    with pytest.raises(SystemExit) as exc:
        mod.main()
    assert exc.value.code == 1
    assert '"status": "ERROR"' in capsys.readouterr().out


def test_node_manifest_builder_success_and_failures(tmp_path):
    script = ROOT / "claude-for-msft-365-install" / "scripts" / "build-manifest.mjs"
    manifest = (
        '<OfficeApp><SourceLocation DefaultValue="https://pivot.claude.ai/taskpane?m=1"/>'
        '<bt:Url id="Taskpane.Url" DefaultValue="https://pivot.claude.ai/taskpane?m=1"/></OfficeApp>'
    )

    class Handler(BaseHTTPRequestHandler):
        def do_GET(self):
            self.send_response(200)
            self.end_headers()
            self.wfile.write(manifest.encode())

        def log_message(self, format, *args):
            return

    server = HTTPServer(("127.0.0.1", 0), Handler)
    thread = threading.Thread(target=server.serve_forever)
    thread.start()
    try:
        port = server.server_port
        env = {**os.environ, "MANIFEST_URL": f"http://127.0.0.1:{port}/manifest.xml"}
        out = tmp_path / "out.xml"
        ok = subprocess.run(
            ["node", str(script), str(out), "gateway_url=https://gateway.test", "gateway_token=secret"],
            env=env,
            text=True,
            capture_output=True,
            check=False,
        )
        assert ok.returncode == 0, ok.stderr
        assert "gateway_url=https%3A%2F%2Fgateway.test&amp;gateway_token=secret" in out.read_text()
        assert "note: gateway_token" in ok.stderr

        bad = subprocess.run(["node", str(script), str(out), "unknown=value"], env=env, text=True, capture_output=True)
        assert bad.returncode == 1
        assert "unknown key" in bad.stderr

        entra = subprocess.run(["node", str(script), str(out), "aws_role_arn=arn:aws:iam::123456789012:role/x"], env=env, text=True, capture_output=True)
        assert entra.returncode == 1
        assert "requires entra_sso=1" in entra.stderr
    finally:
        server.shutdown()
        thread.join(timeout=5)
